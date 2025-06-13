import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import os
import sys
import argparse
import re
from collections import defaultdict
from pathlib import Path
from datetime import datetime

def convert_csv_to_excel(input_folder='.', output_file=None, verbose=False):
    """
    Convert CSV files in a folder to a single Excel workbook with multiple sheets.
    
    Parameters:
    input_folder (str): Folder containing CSV files (default: current directory)
    output_file (str): Output Excel file path (default: auto-generated)
    verbose (bool): Whether to print verbose output
    
    Returns:
    tuple: (success: bool, output_path: str)
    """
    # Generate default output filename if not provided
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'combined_excel_{timestamp}.xlsx'
    
    if verbose:
        print(f"Converting CSV files from {input_folder} to {output_file}")
    
    # Flag to track if any sheet is written
    sheets_written = False
    
    # Create an Excel writer object to write multiple sheets
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Get all CSV files in the input folder
        csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]
        
        if not csv_files:
            print("No CSV files found in the current directory!")
            return False, None
        
        if verbose:
            print(f"Found {len(csv_files)} CSV files")
        
        for filename in csv_files:
            # Construct the file path
            file_path = os.path.join(input_folder, filename)
            
            try:
                # Read the CSV file into a DataFrame
                df = pd.read_csv(file_path)
                # Use the filename (without extension) as the sheet name
                sheet_name = os.path.splitext(filename)[0][:31]  # Sheet name limit is 31 chars
                # Save the DataFrame as a new sheet in the combined Excel file
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                # Mark that a sheet was written
                sheets_written = True
                if verbose:
                    print(f"Added sheet '{sheet_name}' with {len(df)} rows")
            except pd.errors.EmptyDataError:
                print(f"Skipping {filename}: EmptyDataError - No data found in the CSV file.")
            except pd.errors.ParserError:
                print(f"Skipping {filename}: ParserError - Could not parse the CSV file.")
            except Exception as e:
                print(f"Skipping {filename} due to unexpected error: {e}")
        
        # If no sheets were written, create a dummy sheet to avoid an empty workbook error
        if not sheets_written:
            pd.DataFrame({"Message": ["No valid data found"]}).to_excel(writer, sheet_name="NoData")
            return False, None
    
    print(f"Successfully created: {output_file}")
    if verbose:
        print("All valid sheets have been saved into the compiled Excel file")
    return True, output_file

def process_excel_by_object_id(input_path, output_path=None, verbose=False):
    """
    Process the Excel file by separating sheets by object_id and combining related data.
    
    Parameters:
    input_path (str): Path to original input Excel file
    output_path (str): Path for final output file (default: auto-generated)
    verbose (bool): Whether to print verbose output
    
    Returns:
    tuple: (success: bool, output_path: str or None)
    """
    try:
        # Generate default output filename if not provided
        if output_path is None:
            input_stem = Path(input_path).stem
            output_path = f'{input_stem}_formatted.xlsx'
        
        # Get absolute path to see where file is being saved
        output_abs_path = os.path.abspath(output_path)
        
        if verbose:
            print(f"\nProcessing Excel file by object_id: {input_path}")
            print(f"Output will be saved to: {output_abs_path}")
            print("Loading Excel file and processing sheets...")
        
        # Load the Excel file
        excel_data = pd.ExcelFile(input_path)
        
        # Dictionary to store separated data by sheet and object_id
        separated_data = {}
        has_object_id = False
        
        # Step 1: Separate data by object_id in memory
        for sheet_name in excel_data.sheet_names:
            data = pd.read_excel(input_path, sheet_name=sheet_name)
            if verbose:
                print(f"Loaded {len(data)} rows from sheet '{sheet_name}'")
            
            if 'object_id' in data.columns:
                has_object_id = True
                # Group data by 'object_id'
                for object_id, group in data.groupby('object_id'):
                    key = f"{sheet_name}_obj_{object_id}"
                    separated_data[key] = group.copy()
                    if verbose:
                        print(f"Separated {len(group)} rows for '{key}'")
            else:
                if verbose:
                    print(f"'object_id' column not found in sheet '{sheet_name}' - skipping")
        
        if not has_object_id:
            if verbose:
                print("No sheets with 'object_id' column found. No object_id processing needed.")
            return False, None
        
        # Step 2: Group sheets by their stem names and combine
        sheet_groups = defaultdict(list)
        for sheet_key in separated_data.keys():
            # Extract stem name (everything before _obj_)
            stem = re.split(r'_obj_[12]', sheet_key)[0]
            sheet_groups[stem].append(sheet_key)
        
        if verbose:
            print(f"\nCombining {len(sheet_groups)} sheet groups...")
        
        # Step 3: Create final Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for stem, sheet_list in sheet_groups.items():
                # Sort sheets to ensure _obj_1 comes before _obj_2
                sheet_list.sort()
                
                # Start with first sheet data
                combined_df = separated_data[sheet_list[0]].copy()
                
                # Add empty columns and second sheet if it exists
                if len(sheet_list) > 1:
                    # Add 4 empty columns
                    for i in range(4):
                        combined_df[f'empty_{i+1}'] = ''
                    
                    # Get second sheet data and reset its index
                    second_df = separated_data[sheet_list[1]].copy().reset_index(drop=True)
                    
                    # Reset index of first dataframe to ensure proper alignment
                    combined_df = combined_df.reset_index(drop=True)
                    
                    # Concatenate horizontally (side by side) starting from the same row
                    combined_df = pd.concat([combined_df, second_df], axis=1, ignore_index=False)
                
                # Write combined data to new sheet
                combined_df.to_excel(writer, sheet_name=stem, index=False)
                if verbose:
                    print(f"Created combined sheet '{stem}' with {len(combined_df)} rows")
        
        print(f"Object ID processing completed! Output saved to: {output_abs_path}")
        return True, output_path
            
    except Exception as e:
        print(f"An error occurred during object_id processing: {str(e)}")
        return False, None

def calculate_bout_durations(timestamps):
    """
    Calculate bout durations from timestamp pairs.
    Assumes timestamps are in pairs (start, end, start, end, etc.)
    """
    bout_durations = []
    
    # Process pairs of timestamps (0,1 then 2,3 etc.)
    for i in range(0, len(timestamps) - 1, 2):
        if pd.notna(timestamps[i]) and pd.notna(timestamps[i + 1]):
            duration = timestamps[i + 1] - timestamps[i]
            bout_durations.append(duration)
    
    return bout_durations

def process_worksheet_for_di(df, sheet_name):
    """
    Process a single worksheet's data to calculate exploration times and DI.
    Handles both original format and object_id formatted data.
    """
    print(f"Processing sheet for DI: {sheet_name}")
    
    # Check if this is formatted data (has empty columns in middle)
    # Formatted data structure: obj1_data | empty_cols | obj2_data
    
    # Look for the pattern of empty columns that indicate formatted data
    has_empty_columns = False
    empty_col_start = None
    
    # Check for empty columns (typically columns with names like 'empty_1', 'empty_2', etc.)
    if hasattr(df, 'columns'):
        for i, col in enumerate(df.columns):
            if str(col).startswith('empty_') or pd.isna(col) or str(col).strip() == '':
                if empty_col_start is None:
                    empty_col_start = i
                has_empty_columns = True
    
    if has_empty_columns and empty_col_start is not None:
        # This is formatted data - split at the empty columns
        print(f"Detected formatted data structure, splitting at column {empty_col_start}")
        
        # Object 1 data is before the empty columns
        obj1_data = df.iloc[:, :empty_col_start]
        
        # Object 2 data is after the empty columns (skip 4 empty columns)
        obj2_start = empty_col_start + 4
        if obj2_start < df.shape[1]:
            obj2_data = df.iloc[:, obj2_start:]
            
            # Look for timestamp columns in obj2 data
            # Assume first column contains timestamps
            obj2_timestamps = obj2_data.iloc[:, 0].dropna().tolist()
        else:
            obj2_timestamps = []
        
        # Look for timestamp columns in obj1 data  
        # Assume first column contains timestamps
        obj1_timestamps = obj1_data.iloc[:, 0].dropna().tolist()
        
    else:
        # Original format - timestamps in columns A and H
        print("Using original data structure (columns A and H)")
        
        # Extract Object 1 timestamps (Column A, assuming 0-indexed)
        obj1_timestamps = df.iloc[:, 0].dropna().tolist()
        
        # Extract Object 2 timestamps (Column H, assuming 0-indexed = column 7)  
        obj2_timestamps = df.iloc[:, 7].dropna().tolist() if df.shape[1] > 7 else []
    
    # Convert any string timestamps to numeric values
    def clean_timestamps(timestamps):
        cleaned = []
        for ts in timestamps:
            if pd.isna(ts):
                continue
            try:
                # Try to convert to float
                if isinstance(ts, str):
                    # Remove any non-numeric characters except decimal point
                    cleaned_str = ''.join(c for c in ts if c.isdigit() or c == '.')
                    if cleaned_str:
                        cleaned.append(float(cleaned_str))
                else:
                    cleaned.append(float(ts))
            except (ValueError, TypeError):
                print(f"Warning: Could not convert timestamp '{ts}' to numeric value")
                continue
        return cleaned
    
    obj1_timestamps = clean_timestamps(obj1_timestamps)
    obj2_timestamps = clean_timestamps(obj2_timestamps)
    
    # Calculate bout durations
    obj1_durations = calculate_bout_durations(obj1_timestamps)
    obj2_durations = calculate_bout_durations(obj2_timestamps)
    
    # Calculate total exploration times
    obj1_total_time = sum(obj1_durations) if obj1_durations else 0
    obj2_total_time = sum(obj2_durations) if obj2_durations else 0
    total_exploration_time = obj1_total_time + obj2_total_time
    
    # Calculate Discrimination Index (DI)
    if total_exploration_time > 0:
        di = (obj1_total_time - obj2_total_time) / total_exploration_time
    else:
        di = 0
    
    # Round values for display
    obj1_total_time = round(obj1_total_time, 1)
    obj2_total_time = round(obj2_total_time, 1)
    total_exploration_time = round(total_exploration_time, 1)
    di = round(di, 2)
    
    return {
        'Sheet Name': sheet_name,
        'Obj1 Exploration': obj1_total_time,
        'Obj2 Exploration': obj2_total_time,
        'TET': total_exploration_time,
        'DI': di,
        'obj1_durations': obj1_durations,
        'obj2_durations': obj2_durations
    }

def write_di_summary_to_sheet(workbook, sheet_name, summary_data):
    """
    Write DI summary data back to the original sheet with formatting.
    Handles both original and formatted data structures.
    """
    try:
        ws = workbook[sheet_name]
        
        # Find a good location for the summary table
        # Check if this is formatted data by looking for empty columns
        max_col = ws.max_column
        
        # Try to place summary after the main data
        summary_start_col = max_col + 2  # Leave one empty column
        summary_start_row = 10
        
        # Convert column number to letter for Excel
        from openpyxl.utils import get_column_letter
        col_letter_1 = get_column_letter(summary_start_col)
        col_letter_2 = get_column_letter(summary_start_col + 1)
        
        # Add headers and calculations 
        ws[f'{col_letter_1}{summary_start_row}'] = 'Obj1 Exploration'
        ws[f'{col_letter_2}{summary_start_row}'] = summary_data['Obj1 Exploration']
        
        ws[f'{col_letter_1}{summary_start_row + 1}'] = 'Obj2 Exploration'
        ws[f'{col_letter_2}{summary_start_row + 1}'] = summary_data['Obj2 Exploration']
        
        ws[f'{col_letter_1}{summary_start_row + 2}'] = 'TET'
        ws[f'{col_letter_2}{summary_start_row + 2}'] = summary_data['TET']
        
        ws[f'{col_letter_1}{summary_start_row + 3}'] = 'DI'
        ws[f'{col_letter_2}{summary_start_row + 3}'] = summary_data['DI']
        
        # Add bout duration calculations in nearby columns
        bout_col_1 = get_column_letter(summary_start_col + 3)
        bout_col_2 = get_column_letter(summary_start_col + 4)
        
        # Object 1 bout durations
        ws[f'{bout_col_1}1'] = 'Obj1_Explore'
        for i, duration in enumerate(summary_data['obj1_durations']):
            ws[f'{bout_col_1}{i + 3}'] = duration  # Starting from row 3
        
        # Object 2 bout durations  
        ws[f'{bout_col_2}1'] = 'Obj2_Explore'
        for i, duration in enumerate(summary_data['obj2_durations']):
            ws[f'{bout_col_2}{i + 3}'] = duration  # Starting from row 3
        
        # Add summary headers
        ws[f'{col_letter_1}1'] = 'Obj1_Explore_Time'
        ws[f'{col_letter_2}1'] = summary_data['Obj1 Exploration']
        
        ws[f'{bout_col_1}2'] = 'Obj2_Explore_Time'
        ws[f'{bout_col_2}2'] = summary_data['Obj2 Exploration']
        
        # Format the summary table with borders
        summary_range = f'{col_letter_1}{summary_start_row}:{col_letter_2}{summary_start_row + 3}'
        
        # Create border style
        thin_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Apply border to the range
        for row in ws[summary_range]:
            for cell in row:
                cell.border = thin_border
        
    except Exception as e:
        print(f"Error writing DI summary to sheet {sheet_name}: {e}")

def create_di_consolidated_sheet(workbook, all_summaries):
    """
    Create a consolidated summary sheet with all DI results.
    """
    # Remove existing consolidated sheet if it exists
    if "Consolidated Data" in workbook.sheetnames:
        del workbook["Consolidated Data"]
    
    # Create new consolidated sheet
    consolidated_ws = workbook.create_sheet("Consolidated Data")
    
    # Add headers
    consolidated_ws['A1'] = 'Sheet Name'
    consolidated_ws['B1'] = 'Values'
    
    # Format headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    consolidated_ws['A1'].font = header_font
    consolidated_ws['A1'].fill = header_fill
    consolidated_ws['B1'].font = header_font
    consolidated_ws['B1'].fill = header_fill
    
    # Add data from each sheet
    current_row = 2
    
    for summary in all_summaries:
        if summary['Sheet Name'] != 'Consolidated Data':
            # Add sheet name
            consolidated_ws[f'A{current_row}'] = summary['Sheet Name']
            current_row += 1
            
            # Add the 4-row summary
            summary_items = [
                ('Obj1 Exploration', summary['Obj1 Exploration']),
                ('Obj2 Exploration', summary['Obj2 Exploration']),
                ('TET', summary['TET']),
                ('DI', summary['DI'])
            ]
            
            for label, value in summary_items:
                consolidated_ws[f'A{current_row}'] = label
                consolidated_ws[f'B{current_row}'] = value
                current_row += 1
            
            # Add blank row for separation
            current_row += 1
    
    print("DI Consolidated sheet created successfully")

def calculate_discrimination_index(file_path, output_path=None, verbose=False):
    """
    Main function to process an Excel file with exploration behavior data and calculate DI.
    
    Parameters:
    file_path (str): Path to the Excel file to process
    output_path (str): Path for output file (default: auto-generated)
    verbose (bool): Whether to print verbose output
    
    Returns:
    tuple: (success: bool, output_path: str or None, summaries: list)
    """
    try:
        if verbose:
            print(f"\nCalculating Discrimination Index for: {file_path}")
        
        # Generate default output filename if not provided
        if output_path is None:
            input_stem = Path(file_path).stem
            output_path = f'{input_stem}_DI_processed.xlsx'
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Read all sheets except 'Consolidated Data'
        excel_file = pd.ExcelFile(file_path)
        sheet_names = [name for name in excel_file.sheet_names if name != 'Consolidated Data']
        
        all_summaries = []
        
        # Process each worksheet
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                summary = process_worksheet_for_di(df, sheet_name)
                all_summaries.append(summary)
                
                # Write summary back to the original sheet
                write_di_summary_to_sheet(workbook, sheet_name, summary)
                
            except Exception as e:
                print(f"Error processing sheet {sheet_name} for DI: {e}")
                continue
        
        # Create consolidated summary sheet
        if all_summaries:
            create_di_consolidated_sheet(workbook, all_summaries)
        
        # Save the workbook
        workbook.save(output_path)
        workbook.close()
        
        print(f"DI calculation completed. Results saved to: {output_path}")
        
        # Display summary results
        if verbose:
            print("\n--- DISCRIMINATION INDEX RESULTS ---")
            for summary in all_summaries:
                print(f"\nSheet: {summary['Sheet Name']}")
                print(f"  Obj1 Exploration: {summary['Obj1 Exploration']}")
                print(f"  Obj2 Exploration: {summary['Obj2 Exploration']}")
                print(f"  Total Exploration Time (TET): {summary['TET']}")
                print(f"  Discrimination Index (DI): {summary['DI']}")
        
        return True, output_path, all_summaries
        
    except Exception as e:
        print(f"Error calculating DI: {e}")
        return False, None, []

def main():
    parser = argparse.ArgumentParser(
        description="Combined tool: Convert CSV to Excel, format by object_id, and calculate Discrimination Index (DI).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
This combined script handles three main operations:
1. Convert CSV files to Excel workbook
2. Format Excel by separating object_ids into side-by-side columns
3. Calculate Discrimination Index (DI) from exploration behavior data

WORKFLOW MODES:

Mode 1 - Full Pipeline (CSV → Excel → Format → DI):
  %(prog)s --full-pipeline
  %(prog)s --full-pipeline -o final_output.xlsx -v

Mode 2 - CSV to Excel only:
  %(prog)s --csv-only
  %(prog)s --csv-only -o converted.xlsx

Mode 3 - Format existing Excel by object_id:
  %(prog)s --format-only -e existing.xlsx
  %(prog)s --format-only -e existing.xlsx -o formatted.xlsx

Mode 4 - Calculate DI from existing Excel:
  %(prog)s --di-only -e existing.xlsx
  %(prog)s --di-only -e existing.xlsx -o results.xlsx

Mode 5 - Process existing Excel file (Format + DI):
  %(prog)s --process-excel -e existing.xlsx

EXAMPLES:
  %(prog)s --full-pipeline -v                    # Complete workflow with verbose output
  %(prog)s --csv-only -o data.xlsx              # Just convert CSVs
  %(prog)s --di-only -e formatted_data.xlsx     # Just calculate DI from formatted data
  %(prog)s --process-excel -e raw_data.xlsx -v  # Format and calculate DI from existing file
        """
    )
    
    # Mode selection (mutually exclusive group)
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument(
        '--full-pipeline', '--full',
        action='store_true',
        help='Run complete pipeline: CSV→Excel→Format→DI calculation'
    )
    mode_group.add_argument(
        '--csv-only', '--csv',
        action='store_true',
        help='Convert CSV files to Excel only'
    )
    mode_group.add_argument(
        '--format-only', '--format',
        action='store_true',
        help='Format existing Excel file by object_id only'
    )
    mode_group.add_argument(
        '--di-only', '--di',
        action='store_true',
        help='Calculate Discrimination Index from existing Excel file only'
    )
    mode_group.add_argument(
        '--process-excel', '--process',
        action='store_true',
        help='Format existing Excel by object_id AND calculate DI'
    )
    
    # Input options
    parser.add_argument(
        '--excel-file', '-e',
        help='Input Excel file (required for format-only, di-only, and process-excel modes)'
    )
    
    parser.add_argument(
        '--csv-folder', '-c',
        default='.',
        help='Folder containing CSV files (default: current directory)'
    )
    
    # Output options
    parser.add_argument(
        '--output', '-o',
        help='Output file name (default: auto-generated with appropriate suffix)'
    )
    
    # Verbose flag
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Show detailed progress information'
    )
    
    # Parse arguments
    args = parser.parse_args()
    
    # Validate arguments based on mode
    if args.format_only or args.di_only or args.process_excel:
        if not args.excel_file:
            print("Error: --excel-file (-e) is required for format-only, di-only, and process-excel modes!")
            sys.exit(1)
        if not os.path.exists(args.excel_file):
            print(f"Error: Excel file '{args.excel_file}' not found!")
            sys.exit(1)
    
    if args.full_pipeline or args.csv_only:
        if not os.path.exists(args.csv_folder):
            print(f"Error: CSV folder '{args.csv_folder}' not found!")
            sys.exit(1)
        csv_count = len([f for f in os.listdir(args.csv_folder) if f.endswith('.csv')])
        if csv_count == 0:
            print(f"Error: No CSV files found in '{args.csv_folder}'!")
            sys.exit(1)
    
    # Execute based on mode
    try:
        if args.full_pipeline:
            print("=== FULL PIPELINE MODE ===")
            print("Step 1: Converting CSV files to Excel...")
            
            # Step 1: Convert CSV to Excel
            temp_excel = None
            if args.output:
                # Use a temporary name for intermediate file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                temp_excel = f'temp_csv_converted_{timestamp}.xlsx'
            
            success, excel_path = convert_csv_to_excel(args.csv_folder, temp_excel, args.verbose)
            if not success:
                print("CSV conversion failed!")
                sys.exit(1)
            
            current_file = excel_path
            
            # Step 2: Format by object_id (if applicable)
            print("\nStep 2: Checking for object_id formatting...")
            format_success, formatted_path = process_excel_by_object_id(current_file, None, args.verbose)
            
            if format_success:
                # Clean up the previous file if formatting was successful
                if temp_excel and os.path.exists(current_file):
                    os.remove(current_file)
                    if args.verbose:
                        print(f"Removed intermediate file: {current_file}")
                current_file = formatted_path
            else:
                print("No object_id formatting needed, proceeding with current file...")
            
            # Step 3: Calculate DI
            print("\nStep 3: Calculating Discrimination Index...")
            di_success, final_path, summaries = calculate_discrimination_index(current_file, args.output, args.verbose)
            
            if not di_success:
                print("DI calculation failed!")
                sys.exit(1)
            
            # Clean up intermediate file if we have a final output
            if format_success and final_path != current_file and os.path.exists(current_file):
                os.remove(current_file)
                if args.verbose:
                    print(f"Removed intermediate file: {current_file}")
            
            print(f"\n=== FULL PIPELINE COMPLETED ===")
            print(f"Final output: {os.path.abspath(final_path)}")
            
        elif args.csv_only:
            print("=== CSV TO EXCEL MODE ===")
            success, excel_path = convert_csv_to_excel(args.csv_folder, args.output, args.verbose)
            if not success:
                sys.exit(1)
                
        elif args.format_only:
            print("=== FORMAT BY OBJECT_ID MODE ===")
            success, output_path = process_excel_by_object_id(args.excel_file, args.output, args.verbose)
            if not success:
                print("No object_id formatting was needed or formatting failed!")
                sys.exit(1)
                
        elif args.di_only:
            print("=== DISCRIMINATION INDEX CALCULATION MODE ===")
            success, output_path, summaries = calculate_discrimination_index(args.excel_file, args.output, args.verbose)
            if not success:
                sys.exit(1)
                
        elif args.process_excel:
            print("=== PROCESS EXCEL MODE (Format + DI) ===")
            print("Step 1: Formatting by object_id...")
            
            # Step 1: Format by object_id
            format_success, formatted_path = process_excel_by_object_id(args.excel_file, None, args.verbose)
            
            # Determine which file to use for DI calculation
            file_for_di = formatted_path if format_success else args.excel_file
            
            if not format_success:
                print("No object_id formatting needed, proceeding with DI calculation...")
            
            # Step 2: Calculate DI
            print("\nStep 2: Calculating Discrimination Index...")
            di_success, final_path, summaries = calculate_discrimination_index(file_for_di, args.output, args.verbose)
            
            if not di_success:
                print("DI calculation failed!")
                sys.exit(1)
            
            # Clean up intermediate file if needed
            if format_success and final_path != formatted_path and os.path.exists(formatted_path):
                os.remove(formatted_path)
                if args.verbose:
                    print(f"Removed intermediate file: {formatted_path}")
            
            print(f"\n=== EXCEL PROCESSING COMPLETED ===")
            print(f"Final output: {os.path.abspath(final_path)}")
        
        print("\n✅ All operations completed successfully!")
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Operation interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()

"""
COMBINED SCRIPT PURPOSE:
This script combines CSV-to-Excel conversion, object_id formatting, and Discrimination Index calculation
into a single comprehensive tool for behavioral data analysis.

CAPABILITIES:
1. CSV to Excel Conversion: Combines multiple CSV files into a single Excel workbook
2. Object ID Formatting: Separates data by object_id and arranges side-by-side
3. Discrimination Index Calculation: Analyzes exploration behavior and calculates DI

WORKFLOW MODES:
- Full Pipeline: CSV → Excel → Format → DI (complete end-to-end processing)
- Individual Steps: Run any step independently
- Process Excel: Format + DI calculation for existing Excel files


DATA ASSUMPTIONS FOR DI CALCULATION:
- Column A (index 0): Object 1 timestamps (pairs: start/end, start/end, etc.)
- Column H (index 7): Object 2 timestamps (pairs: start/end, start/end, etc.)
- Timestamps in consecutive rows forming pairs (row 1&2 = bout 1, row 3&4 = bout 2, etc.)

DISCRIMINATION INDEX (DI):
- Formula: (Object1_Time - Object2_Time) / Total_Exploration_Time
- Range: -1 to +1 (positive = preference for Object 1, negative = preference for Object 2)
"""